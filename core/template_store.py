"""
core/template_store.py

Stores reusable Excel export templates for custom column layouts.
"""

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parent.parent
TEMPLATE_DIR = ROOT_DIR / "templates"

DEFAULT_TEMPLATE = {
    "name": "Default",
    "columns": [
        {"key": "row_number", "header": "#"},
        {"key": "date", "header": "Date"},
        {"key": "category", "header": "Category"},
        {"key": "particular", "header": "Particular"},
        {"key": "amount", "header": "Amount"},
        {"key": "remarks", "header": "Remarks"},
        {"key": "working", "header": "Working"},
    ],
}


def _ensure_dir() -> Path:
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    return TEMPLATE_DIR


def list_templates() -> List[str]:
    _ensure_dir()
    return sorted([path.stem for path in TEMPLATE_DIR.glob("*.json") if path.is_file()])


def load_template(name: str) -> Optional[Dict[str, Any]]:
    if not name or name == "Default":
        return DEFAULT_TEMPLATE

    _ensure_dir()
    path = TEMPLATE_DIR / f"{name}.json"
    if not path.exists():
        return None

    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def save_template(name: str, headers: List[str]) -> Dict[str, Any]:
    _ensure_dir()
    clean_name = re.sub(r"[^A-Za-z0-9 _.-]+", "", name).strip() or "custom_template"
    columns = []
    for header in headers:
        if header is None:
            continue
        columns.append({"key": _infer_key(str(header)), "header": str(header)})

    template = {"name": clean_name, "columns": columns}
    path = TEMPLATE_DIR / f"{clean_name}.json"
    with path.open("w", encoding="utf-8") as handle:
        json.dump(template, handle, indent=2)
    return template


def delete_template(name: str) -> bool:
    if not name or name == "Default":
        return False

    _ensure_dir()
    path = TEMPLATE_DIR / f"{name}.json"
    if path.exists():
        path.unlink()
        return True
    return False


def read_headers_from_excel(file_path: str) -> List[str]:
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active
    if sheet.max_row == 0 or sheet.max_column == 0:
        return []

    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = []
    for value in first_row:
        headers.append("" if value is None else str(value))
    return headers


def _infer_key(header: str) -> str:
    normalized = header.strip().lower()
    if "date" in normalized:
        return "date"
    if "category" in normalized or "type" in normalized or "account" in normalized:
        return "category"
    if "particular" in normalized or "description" in normalized or "counter" in normalized or "name" in normalized:
        return "particular"
    if "amount" in normalized or "total" in normalized or "value" in normalized or "sum" in normalized:
        return "amount"
    if "remark" in normalized or "note" in normalized or "memo" in normalized:
        return "remarks"
    if "working" in normalized:
        return "working"
    if normalized in {"#", "no", "number", "row"} or "row" in normalized:
        return "row_number"
    return "remarks"
