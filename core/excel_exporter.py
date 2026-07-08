"""
core/excel_exporter.py

Takes a list of Transaction objects and builds a .xlsx file using
whatever columns and titles the user chose in the app -- no hardcoded
column names. See app.py's "Customize your columns" section for how
the column list gets built.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Every field we know how to pull from a parsed transaction, plus a
# sensible default column width for each.
FIELD_WIDTHS = {
    "date": 12,
    "category": 12,
    "particular": 28,
    "amount": 14,
    "remarks": 20,
    "transaction_id": 22,
    "source_file": 30,
}


def build_workbook(transactions: list, columns: list) -> Workbook:
    """
    transactions: list of Transaction objects (see models/transaction.py)
    columns: list of {"field": <str>, "title": <str>} in the exact
             order and with the exact titles the user chose. "field"
             must be one of: date, category, particular, amount,
             remarks, transaction_id, source_file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # --- Header row (using the user's own custom titles) ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col["title"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # --- Data rows ---
    for i, txn in enumerate(transactions):
        row_idx = i + 2  # row 1 is the header
        field_values = {
            "date": txn.date or "",
            "category": txn.category or "",
            "particular": txn.particular or "",
            "amount": txn.amount if txn.amount is not None else "",
            "remarks": txn.remarks or "",
            "transaction_id": txn.transaction_id or "",
            "source_file": txn.source_file or "",
        }
        for col_idx, col in enumerate(columns, start=1):
            value = field_values.get(col["field"], "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    # --- Column widths ---
    for col_idx, col in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = FIELD_WIDTHS.get(col["field"], 16)

    ws.freeze_panes = "A2"
    return wb